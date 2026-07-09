# import openpyxl

# print(openpyxl.__version__)



from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Name"
ws["B1"] = "Marks"

ws["A2"] = "Vennela"
ws["B2"] = 95

wb.save("student.xlsx")

print("Excel file created successfully!")